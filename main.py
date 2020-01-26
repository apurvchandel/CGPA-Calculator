# import numpy as nppip
import pandas as pd
from openpyxl import Workbook, load_workbook,worksheet
from openpyxl.utils.dataframe import dataframe_to_rows


sheetlist =["MATHS","OOPS","CSO","MUI","DSA"]
j=0
wb = Workbook()
ws = wb.active
wb.save('filename.xlsx')
for j in range(0,5):
    data = pd.read_excel("input.xlsx", sheetname=sheetlist[j], index_col=0)
    wb = load_workbook('filename.xlsx')
    sheet1 = wb.create_sheet(sheetlist[j],-1)

    active = wb[sheetlist[j]]

    for x in dataframe_to_rows(data):
        active.append(x)

    sheet1.delete_rows(2,amount =1)
    sheet1.delete_cols(1,amount =1)
    wb.save('filename.xlsx')
    wb = load_workbook('filename.xlsx')
    ws = wb[sheetlist[j]]
    l = ws.max_row 

    ws['I1']="Total"

    for i in range(2,l+1):
        if ws['C'+str(i)].value == 'Ab':
            ws['C'+str(i)] = 0
        if ws['D'+str(i)].value == 'Ab':
            ws['D'+str(i)] = 0
        if ws['E'+str(i)].value == 'Ab':
            ws['E'+str(i)] = 0
        if ws['F'+str(i)].value == 'Ab':
            ws['F'+str(i)] = 0
        if ws['G'+str(i)].value == 'Ab':
            ws['G'+str(i)] = 0
        ws['I'+str(i)] = (float(ws['C'+str(i)].value)+float(ws['D'+str(i)].value)+float(ws['E'+str(i)].value)+float(ws['F'+str(i)].value))*0.75+(float(ws['G'+str(i)].value)*0.25)

    ws.delete_cols(3,amount =1)
    ws.delete_cols(3,amount =1)
    ws.delete_cols(3,amount =1)
    ws.delete_cols(3,amount =1)
    ws.delete_cols(3,amount =1)
    ws.delete_cols(3,amount =1)

    ws['D1']="Grade"
    wb.save('filename.xlsx')
    wb = load_workbook('filename.xlsx')
    ws = wb[sheetlist[j]]

    df = pd.read_excel('filename.xlsx')
    ma = df["Total"].max()

    for i in range(2,l+1):
        ws['E'+str(i)] = (float(ws['C'+str(i)].value) *100)/ma
    for i in range(2,l+1):
        if float(ws['E'+str(i)].value) <100.01 and float(ws['E'+str(i)].value) >= 95:
             ws['D'+str(i)] = "AA"    
        elif float(ws['E'+str(i)].value) < 95.01 and float(ws['E'+str(i)].value) >= 85:
             ws['D'+str(i)] = "AB" 
        elif float(ws['E'+str(i)].value) < 85.01 and float(ws['E'+str(i)].value) >= 75:
             ws['D'+str(i)] = "BB"
        elif float(ws['E'+str(i)].value) < 75.01 and float(ws['E'+str(i)].value) >= 65:
             ws['D'+str(i)] = "BC"
        elif float(ws['E'+str(i)].value) < 65.01 and float(ws['E'+str(i)].value) >= 55:
             ws['D'+str(i)] = "CC"
        elif float(ws['E'+str(i)].value) < 55.01 and float(ws['E'+str(i)].value) >= 45:
             ws['D'+str(i)] = "CD"
        elif float(ws['E'+str(i)].value) < 45.01 and float(ws['E'+str(i)].value) >= 35:
             ws['D'+str(i)] = "DD"
        else:
            ws['D'+str(i)] = "FF"

    ws.delete_cols(5,amount =1)               
    wb.save('filename.xlsx')

std = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(std)    
wb.save('filename.xlsx')

data = pd.read_excel("input.xlsx", sheetname=1, index_col=0)
wb = load_workbook('filename.xlsx')
sheet1 = wb.create_sheet("Final")
active = wb["Final"]
ws1 =active
for x in dataframe_to_rows(data):
    active.append(x)

sheet1.delete_rows(2,amount =1)
sheet1.delete_cols(1,amount =1)

ws1.delete_cols(3,amount =1)
ws1.delete_cols(3,amount =1)
ws1.delete_cols(3,amount =1)
ws1.delete_cols(3,amount =1)
ws1.delete_cols(3,amount =1)
ws1.delete_cols(3,amount =1)

ws1['C1'] = "MATHS"
ws1['D1'] = "OOPS"
ws1['E1'] = "CSO"
ws1['F1'] = "MUI"
ws1['G1'] = "DSA"
ws1['H1'] = "SGPA"

x=["C","D","E","F","G"]
for j in range(0,5):
    ws = wb[sheetlist[j]]
    for k in range(2,l+1):
        ws1[x[j]+str(k)] = ws['D'+str(k)].value

ws2 =wb["Final"]
y ={"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"FF":0}
z=["C","D","E","F","G"]
for k in range(2,l+1):
    t=0.0
    for b in range(0,5):
        t=t+y[ws2[z[b]+str(k)].value]
    ws2['H'+str(k)] = t/5
    wb.save('filename.xlsx')

wb.save('filename.xlsx')
