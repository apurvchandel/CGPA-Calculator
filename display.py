# 
f=open("user.txt","r")
k=f.read()
k=k.upper()
f.close()
from openpyxl import Workbook,load_workbook

wb = Workbook()
wb = load_workbook('filename.xlsx')
ws = wb["Final"]
l = ws.max_row

a=[]
x = ["B","C","D","E","F","G","H"]
for m in range(2,l+1):
    if ws["A"+str(m)].value == k:
        ws["A"+str(m)]
        for n in range(0,7):
            a.append(ws[x[n]+str(m)].value)
        break
           
# file_name =  "input.xlsx"
from tkinter import *
screen = Tk()
screen.geometry("550x550")
screen.title("Mark Sheet")
heading = Label(text = "Mark Sheet", bg = "grey", fg = "black", width = "500", height = "3")
heading.pack()
screen.wm_iconbitmap('IIIT-N.ico')
firstname_text = Label(text = "Name :  ",)
firstname_text.place(x = 20, y = 100)
name=Label(text=str(a[0]),)
name.place(x=300,y=100)
firstname_text = Label(text = "Enrollment Number : ",)
firstname_text.place(x = 20, y = 150)
name=Label(text=k,)
name.place(x=300,y=150)
firstname_text = Label(text = "Maths : ",)
firstname_text.place(x = 20, y = 200)
name=Label(text=a[1],)
name.place(x=300,y=200)
firstname_text = Label(text = "OOPS ",)
firstname_text.place(x = 20, y = 250)
name=Label(text=a[2],)
name.place(x=300,y=250)
firstname_text = Label(text = "CSO ",)
firstname_text.place(x = 20, y = 300)
name=Label(text=a[3],)
name.place(x=300,y=300)
firstname_text = Label(text = "MUI ",)
firstname_text.place(x = 20, y = 350)
name=Label(text=a[4],)
name.place(x=300,y=350)
firstname_text = Label(text = "DSA ",)
firstname_text.place(x = 20, y = 400)
name=Label(text=a[5],)
name.place(x=300,y=400)
firstname_text = Label(text = "SGPA ",)
firstname_text.place(x = 20, y = 450)
name=Label(text=a[6],)
name.place(x=300,y=450)

mainloop()